from __future__ import annotations

import calendar
import os
import sqlite3
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from functools import wraps
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Optional, TypeVar

from flask import (
    Flask,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "hotel_casa_bella.db"

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")

EXTRA_PERSON_FEE = 75.0

INDIVIDUAL_EXTRA_ALLOWED = {
    "Habitación 201", "Habitación 202", "Habitación 203",
    "Habitación 301", "Habitación 302", "Habitación 303",
    "Habitación 401", "Habitación 402", "Habitación 403",
}

ROOM_PRICE_BY_NAME: dict[str, float] = {
    "Habitación 201": 300.0,
    "Habitación 202": 300.0,
    "Habitación 203": 300.0,
    "Habitación 301": 300.0,
    "Habitación 302": 300.0,
    "Habitación 303": 300.0,
    "Habitación 401": 300.0,
    "Habitación 402": 300.0,
    "Habitación 403": 300.0,
    "Habitación 208": 200.0,
    "Habitación 308": 200.0,
    "Habitación 204": 550.0,
    "Habitación 304": 550.0,
    "Habitación 205": 450.0,
    "Habitación 206": 450.0,
    "Habitación 305": 450.0,
    "Habitación 306": 450.0,
    "Habitación 404": 450.0,
    "Habitación 405": 450.0,
    "Habitación 207": 500.0,
    "Habitación 307": 500.0,
    "Habitación 407": 500.0,
    "VIP Personal 1": 400.0,
    "VIP Personal 2": 400.0,
    "VIP Personal 3": 650.0,
    "Suite Presidencial": 650.0,
}

ROOM_CATEGORY_BY_NAME: dict[str, str] = {
    "Habitación 208": "Individual Simple",
    "Habitación 308": "Individual Simple",
    "Habitación 201": "Individual Delux",
    "Habitación 202": "Individual Delux",
    "Habitación 203": "Individual Delux",
    "Habitación 301": "Individual Delux",
    "Habitación 302": "Individual Delux",
    "Habitación 303": "Individual Delux",
    "Habitación 401": "Individual Delux",
    "Habitación 402": "Individual Delux",
    "Habitación 403": "Individual Delux",
    "Habitación 205": "Dobles",
    "Habitación 206": "Dobles",
    "Habitación 305": "Dobles",
    "Habitación 306": "Dobles",
    "Habitación 404": "Dobles",
    "Habitación 405": "Dobles",
    "Habitación 204": "Triples",
    "Habitación 304": "Triples",
    "Habitación 207": "Premium VIP",
    "Habitación 307": "Premium VIP",
    "Habitación 407": "Premium VIP",
    "VIP Personal 1": "VIP Personal",
    "VIP Personal 2": "VIP Personal",
    "VIP Personal 3": "Presidencial",
    "Suite Presidencial": "Presidencial",
}

ROOM_CAPACITY_BY_NAME: dict[str, int] = {
    "Habitación 208": 1,
    "Habitación 308": 1,
    "Habitación 201": 1,
    "Habitación 202": 1,
    "Habitación 203": 1,
    "Habitación 301": 1,
    "Habitación 302": 1,
    "Habitación 303": 1,
    "Habitación 401": 1,
    "Habitación 402": 1,
    "Habitación 403": 1,
    "Habitación 205": 3,
    "Habitación 206": 3,
    "Habitación 305": 3,
    "Habitación 306": 3,
    "Habitación 404": 3,
    "Habitación 405": 3,
    "Habitación 204": 4,
    "Habitación 304": 4,
    "Habitación 207": 2,
    "Habitación 307": 2,
    "Habitación 407": 2,
    "VIP Personal 1": 1,
    "VIP Personal 2": 1,
    "VIP Personal 3": 2,
    "Suite Presidencial": 2,
}

OFFICIAL_ROOM_NAMES = set(ROOM_PRICE_BY_NAME.keys()) | {"Sala 408"}

F = TypeVar("F", bound=Callable[..., Any])


# ==========================================================
# DB
# ==========================================================
def get_db() -> sqlite3.Connection:
    if "db" not in g:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(_: Any) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    schema_path = BASE_DIR / "schema.sql"
    with closing(sqlite3.connect(DB_PATH)) as conn:
        with open(schema_path, "r", encoding="utf-8") as f:
            conn.executescript(f.read())
        conn.commit()


def ensure_hotel_checkins_table() -> None:
    if not DB_PATH.exists():
        return

    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.row_factory = sqlite3.Row

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS hotel_checkins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reservation_id INTEGER,
                source_type TEXT NOT NULL DEFAULT 'direct'
                    CHECK (source_type IN ('direct', 'reservation')),

                guest_name TEXT NOT NULL,
                phone TEXT NOT NULL,
                room_id INTEGER NOT NULL,
                check_in TEXT NOT NULL,
                check_out TEXT NOT NULL,
                guests INTEGER NOT NULL,
                extra_person INTEGER NOT NULL DEFAULT 0,
                special_requests TEXT,

                payment_method TEXT NOT NULL DEFAULT 'cash'
                    CHECK (payment_method IN ('cash', 'card', 'transfer', 'mixed')),
                total_amount REAL NOT NULL DEFAULT 0,
                amount_paid REAL NOT NULL DEFAULT 0,
                amount_pending REAL NOT NULL DEFAULT 0,

                status TEXT NOT NULL DEFAULT 'checked_in'
                    CHECK (status IN ('checked_in', 'checked_out', 'cancelled')),

                created_by INTEGER NOT NULL,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                checked_out_at TEXT,

                FOREIGN KEY (reservation_id) REFERENCES reservations(id) ON DELETE SET NULL,
                FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
                FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE CASCADE
            )
            """
        )

        cols = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(hotel_checkins)").fetchall()
        }

        if "reservation_id" not in cols:
            conn.execute("ALTER TABLE hotel_checkins ADD COLUMN reservation_id INTEGER")

        if "source_type" not in cols:
            conn.execute("ALTER TABLE hotel_checkins ADD COLUMN source_type TEXT NOT NULL DEFAULT 'direct'")

        if "payment_method" not in cols:
            conn.execute("ALTER TABLE hotel_checkins ADD COLUMN payment_method TEXT NOT NULL DEFAULT 'cash'")

        if "total_amount" not in cols:
            conn.execute("ALTER TABLE hotel_checkins ADD COLUMN total_amount REAL NOT NULL DEFAULT 0")

        if "amount_paid" not in cols:
            conn.execute("ALTER TABLE hotel_checkins ADD COLUMN amount_paid REAL NOT NULL DEFAULT 0")

        if "amount_pending" not in cols:
            conn.execute("ALTER TABLE hotel_checkins ADD COLUMN amount_pending REAL NOT NULL DEFAULT 0")

        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_hotel_checkins_room_id ON hotel_checkins(room_id)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_hotel_checkins_status ON hotel_checkins(status)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_hotel_checkins_check_in ON hotel_checkins(check_in)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_hotel_checkins_check_out ON hotel_checkins(check_out)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_hotel_checkins_reservation_id ON hotel_checkins(reservation_id)"
        )

        conn.commit()


def ensure_reservations_alert_column() -> None:
    if not DB_PATH.exists():
        return
    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.row_factory = sqlite3.Row
        cols = conn.execute("PRAGMA table_info(reservations)").fetchall()
        col_names = {c["name"] for c in cols}
        if "admin_alert_dismissed" not in col_names:
            conn.execute(
                "ALTER TABLE reservations ADD COLUMN admin_alert_dismissed INTEGER NOT NULL DEFAULT 0"
            )
            conn.commit()


def migrate_users_table_for_superadmin() -> None:
    if not DB_PATH.exists():
        return

    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.row_factory = sqlite3.Row
        create_sql_row = conn.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='users'"
        ).fetchone()

        if not create_sql_row or not create_sql_row["sql"]:
            return

        if "superadmin" in create_sql_row["sql"].lower():
            return

        conn.execute("PRAGMA foreign_keys = OFF")
        conn.execute("BEGIN TRANSACTION")
        try:
            conn.execute("ALTER TABLE users RENAME TO users_old")
            conn.execute(
                """
                CREATE TABLE users (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  full_name TEXT NOT NULL,
                  email TEXT NOT NULL UNIQUE,
                  password_hash TEXT NOT NULL,
                  role TEXT NOT NULL CHECK(role IN ('admin','client','superadmin')),
                  created_at TEXT NOT NULL DEFAULT (datetime('now'))
                )
                """
            )
            conn.execute(
                """
                INSERT INTO users (id, full_name, email, password_hash, role, created_at)
                SELECT id, full_name, email, password_hash, role, created_at
                FROM users_old
                """
            )
            conn.execute("DROP TABLE users_old")
            conn.execute("COMMIT")
        except Exception:
            conn.execute("ROLLBACK")
            raise
        finally:
            conn.execute("PRAGMA foreign_keys = ON")


def ensure_role_indexes_and_superadmin() -> None:
    if not DB_PATH.exists():
        return

    migrate_users_table_for_superadmin()

    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute("CREATE INDEX IF NOT EXISTS idx_users_role ON users(role)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_reservations_created_at ON reservations(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_reservations_status ON reservations(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_reservations_user_id ON reservations(user_id)")

        superadmin_email = os.environ.get("SUPERADMIN_EMAIL", "superadmin@casabella.com")
        superadmin_pass = os.environ.get("SUPERADMIN_PASSWORD", "SuperAdmin123*")
        admin_email = os.environ.get("ADMIN_EMAIL", "admin@casabella.com")
        admin_pass = os.environ.get("ADMIN_PASSWORD", "Admin123*")

        exists = conn.execute("SELECT id FROM users WHERE email = ?", (superadmin_email,)).fetchone()
        if exists is None:
            conn.execute(
                """
                INSERT INTO users (full_name, email, password_hash, role)
                VALUES (?, ?, ?, 'superadmin')
                """,
                ("Super Administrador", superadmin_email, generate_password_hash(superadmin_pass)),
            )

        exists_admin = conn.execute("SELECT id FROM users WHERE email = ?", (admin_email,)).fetchone()
        if exists_admin is None:
            conn.execute(
                """
                INSERT INTO users (full_name, email, password_hash, role)
                VALUES (?, ?, ?, 'admin')
                """,
                ("Administrador", admin_email, generate_password_hash(admin_pass)),
            )

        conn.commit()


# ==========================================================
# AUTH
# ==========================================================
def current_user() -> Optional[sqlite3.Row]:
    uid = session.get("user_id")
    if not uid:
        return None
    db = get_db()
    return db.execute("SELECT * FROM users WHERE id = ?", (uid,)).fetchone()


def user_has_role(user: Optional[sqlite3.Row], *allowed_roles: str) -> bool:
    if user is None:
        return False
    hierarchy = {"client": 1, "admin": 2, "superadmin": 3}
    role = user["role"]
    user_level = hierarchy.get(role, 0)
    return any(user_level >= hierarchy.get(r, 999) for r in allowed_roles)


def login_required(view: F) -> F:
    @wraps(view)
    def wrapped(*args: Any, **kwargs: Any):
        if not session.get("user_id"):
            flash("Debes iniciar sesión.", "error")
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped  # type: ignore


def role_required(*roles: str):
    def decorator(view: F) -> F:
        @wraps(view)
        def wrapped(*args: Any, **kwargs: Any):
            user = current_user()
            if user is None:
                flash("Debes iniciar sesión.", "error")
                return redirect(url_for("login"))
            if not user_has_role(user, *roles):
                flash("No tienes permisos para acceder aquí.", "error")
                return redirect(url_for("my_reservations"))
            return view(*args, **kwargs)
        return wrapped  # type: ignore
    return decorator


admin_required = role_required("admin")
superadmin_required = role_required("superadmin")


@app.context_processor
def inject_user():
    user = current_user()
    return {
        "current_user": user,
        "is_admin": user_has_role(user, "admin"),
        "is_superadmin": user_has_role(user, "superadmin"),
    }


# ==========================================================
# HELPERS HOTEL
# ==========================================================
def apply_room_updates() -> dict[str, int]:
    if not DB_PATH.exists():
        return {"updated": 0, "renamed": 0}

    renamed = 0
    updated = 0
    rename_map = {
        "Casa 1": "VIP Personal 1",
        "Casa 2": "VIP Personal 2",
        "Casa 3": "VIP Personal 3",
        "Suite": "Suite Presidencial",
    }

    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.row_factory = sqlite3.Row
        rooms = conn.execute("SELECT id, name FROM rooms").fetchall()
        for r in rooms:
            rid = r["id"]
            name = r["name"]

            new_name = rename_map.get(name, name)
            if new_name != name:
                conn.execute("UPDATE rooms SET name = ? WHERE id = ?", (new_name, rid))
                name = new_name
                renamed += 1

            if name in ROOM_PRICE_BY_NAME:
                conn.execute("UPDATE rooms SET price_per_night = ? WHERE id = ?", (float(ROOM_PRICE_BY_NAME[name]), rid))
                updated += 1
            if name in ROOM_CATEGORY_BY_NAME:
                conn.execute("UPDATE rooms SET category = ? WHERE id = ?", (ROOM_CATEGORY_BY_NAME[name], rid))
            if name in ROOM_CAPACITY_BY_NAME:
                conn.execute("UPDATE rooms SET capacity = ? WHERE id = ?", (int(ROOM_CAPACITY_BY_NAME[name]), rid))
        conn.commit()

    return {"updated": updated, "renamed": renamed}


def hide_non_official_rooms() -> int:
    if not DB_PATH.exists():
        return 0
    hidden = 0
    with closing(sqlite3.connect(DB_PATH)) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT id, name FROM rooms").fetchall()
        for r in rows:
            if r["name"] not in OFFICIAL_ROOM_NAMES:
                conn.execute("UPDATE rooms SET capacity = 0 WHERE id = ?", (r["id"],))
                hidden += 1
        conn.commit()
    return hidden


def get_pending_admin_alerts(db: sqlite3.Connection):
    return db.execute(
        """
        SELECT r.*, rm.name AS room_name, u.full_name AS user_name
        FROM reservations r
        JOIN rooms rm ON rm.id = r.room_id
        JOIN users u ON u.id = r.user_id
        WHERE r.admin_alert_dismissed = 0
          AND r.status != 'checked_out'
        ORDER BY r.created_at DESC
        """
    ).fetchall()


@dataclass
class ReservationData:
    guest_name: str
    email: str
    phone: str
    room_id: int
    check_in: date
    check_out: date
    guests: int
    extra_person: bool
    special_requests: str


def parse_reservation_form(form: Any) -> ReservationData:
    check_in = datetime.strptime(form["check_in"], "%Y-%m-%d").date()
    check_out = datetime.strptime(form["check_out"], "%Y-%m-%d").date()
    return ReservationData(
        guest_name=(form.get("guest_name") or "").strip(),
        email=(form.get("email") or "").strip(),
        phone=(form.get("phone") or "").strip(),
        room_id=int(form["room_id"]),
        check_in=check_in,
        check_out=check_out,
        guests=int(form["guests"]),
        extra_person=(form.get("extra_person") == "1"),
        special_requests=(form.get("special_requests") or "").strip(),
    )


def get_room_meta(db: sqlite3.Connection):
    rooms = db.execute(
        """
        SELECT id, name, category, price_per_night, capacity
        FROM rooms
        WHERE capacity > 0
        ORDER BY category, name
        """
    ).fetchall()

    rooms_meta = [
        {
            "id": r["id"],
            "name": r["name"],
            "category": r["category"],
            "price": float(r["price_per_night"]),
            "capacity": int(r["capacity"]),
            "extraAllowed": (r["name"] in INDIVIDUAL_EXTRA_ALLOWED),
        }
        for r in rooms
    ]
    categories = sorted({rm["category"] for rm in rooms_meta if rm.get("category")})
    return rooms_meta, categories


def reservation_has_extra_person_from_text(text: str | None) -> bool:
    special = (text or "").lower()
    return "persona extra" in special


def calculate_stay_total(room_price: float, check_in_str: str, check_out_str: str, extra_person: bool) -> float:
    check_in = datetime.strptime(check_in_str, "%Y-%m-%d").date()
    check_out = datetime.strptime(check_out_str, "%Y-%m-%d").date()
    nights = max((check_out - check_in).days, 1)
    base = float(room_price) * nights
    extra = EXTRA_PERSON_FEE * nights if extra_person else 0.0
    return round(base + extra, 2)


def calculate_reservation_total(reservation_row: sqlite3.Row) -> float:
    return calculate_stay_total(
        float(reservation_row["price_per_night"]),
        reservation_row["check_in"],
        reservation_row["check_out"],
        reservation_has_extra_person_from_text(reservation_row["special_requests"]),
    )


def validate_capacity(data: ReservationData, room: sqlite3.Row, errors: list[str]) -> None:
    room_name = room["name"]
    base_capacity = int(room["capacity"])

    if room_name in INDIVIDUAL_EXTRA_ALLOWED:
        max_allowed = base_capacity + 1
        if data.guests == 2 and not data.extra_person:
            errors.append("Si van 2 personas en habitación Individual, debes marcar 'Persona extra'.")
        if data.extra_person and data.guests != 2:
            errors.append("La opción 'Persona extra' solo aplica cuando van 2 personas.")
        if data.guests > max_allowed:
            errors.append(f"{room_name} permite máximo {max_allowed} huéspedes.")
    else:
        if data.extra_person:
            errors.append("Esta habitación no permite persona extra.")
        if data.guests > base_capacity:
            errors.append(f"{room_name} permite máximo {base_capacity} huéspedes.")


def validate_reservation(data: ReservationData, db: sqlite3.Connection) -> list[str]:
    errors: list[str] = []

    if not data.guest_name:
        errors.append("El nombre del huésped es obligatorio.")
    if not data.phone:
        errors.append("El teléfono es obligatorio.")
    if data.check_in < date.today():
        errors.append("La fecha de llegada no puede ser anterior a hoy.")
    if data.check_out <= data.check_in:
        errors.append("La fecha de salida debe ser posterior a la fecha de llegada.")
    if data.guests < 1:
        errors.append("Debe registrar al menos 1 huésped.")

    room = db.execute("SELECT * FROM rooms WHERE id = ?", (data.room_id,)).fetchone()
    if room is None:
        errors.append("La habitación seleccionada no existe.")
        return errors

    validate_capacity(data, room, errors)

    overlapping = db.execute(
        """
        SELECT id
        FROM reservations
        WHERE room_id = ?
          AND status IN ('confirmed', 'checked_in')
          AND NOT (check_out <= ? OR check_in >= ?)
        LIMIT 1
        """,
        (data.room_id, data.check_in.isoformat(), data.check_out.isoformat()),
    ).fetchone()

    if overlapping is not None:
        errors.append("Habitación no disponible para esta fecha.")

    return errors


def validate_direct_checkin(data: ReservationData, db: sqlite3.Connection) -> list[str]:
    errors: list[str] = []

    if not data.guest_name:
        errors.append("El nombre del huésped es obligatorio.")
    if not data.phone:
        errors.append("El teléfono es obligatorio.")
    if data.check_in < date.today():
        errors.append("La fecha de ingreso no puede ser anterior a hoy.")
    if data.check_out <= data.check_in:
        errors.append("La fecha de salida debe ser posterior a la fecha de ingreso.")
    if data.guests < 1:
        errors.append("Debe registrar al menos 1 huésped.")

    room = db.execute("SELECT * FROM rooms WHERE id = ?", (data.room_id,)).fetchone()
    if room is None:
        errors.append("La habitación seleccionada no existe.")
        return errors

    validate_capacity(data, room, errors)

    overlapping_reservation = db.execute(
        """
        SELECT id
        FROM reservations
        WHERE room_id = ?
          AND status IN ('confirmed', 'checked_in')
          AND NOT (check_out <= ? OR check_in >= ?)
        LIMIT 1
        """,
        (data.room_id, data.check_in.isoformat(), data.check_out.isoformat()),
    ).fetchone()

    if overlapping_reservation is not None:
        errors.append("La habitación ya está reservada para esas fechas.")
        return errors

    overlapping_checkin = db.execute(
        """
        SELECT id
        FROM hotel_checkins
        WHERE room_id = ?
          AND status = 'checked_in'
          AND NOT (check_out <= ? OR check_in >= ?)
        LIMIT 1
        """,
        (data.room_id, data.check_in.isoformat(), data.check_out.isoformat()),
    ).fetchone()

    if overlapping_checkin is not None:
        errors.append("La habitación ya está ocupada por un huésped hospedado.")
        return errors

    return errors


def reservation_stats(db: sqlite3.Connection) -> dict[str, Any]:
    today = date.today().isoformat()
    total_rooms = db.execute("SELECT COUNT(*) AS count FROM rooms WHERE capacity > 0").fetchone()["count"]

    active_reservations = db.execute(
        """
        SELECT COUNT(*) AS count
        FROM reservations
        WHERE status IN ('confirmed', 'checked_in')
          AND check_out > ?
        """,
        (today,),
    ).fetchone()["count"]

    active_checkins = db.execute(
        """
        SELECT COUNT(*) AS count
        FROM hotel_checkins
        WHERE status = 'checked_in'
          AND check_out > ?
        """,
        (today,),
    ).fetchone()["count"]

    occupancy = round(((active_reservations + active_checkins) / total_rooms) * 100, 1) if total_rooms else 0

    upcoming_checkins = db.execute(
        """
        SELECT r.*, rm.name AS room_name
        FROM reservations r
        JOIN rooms rm ON rm.id = r.room_id
        WHERE r.check_in >= ?
        ORDER BY r.check_in ASC
        LIMIT 5
        """,
        (today,),
    ).fetchall()

    total_users = db.execute("SELECT COUNT(*) AS count FROM users").fetchone()["count"]
    total_clients = db.execute("SELECT COUNT(*) AS count FROM users WHERE role = 'client'").fetchone()["count"]
    total_admins = db.execute("SELECT COUNT(*) AS count FROM users WHERE role = 'admin'").fetchone()["count"]
    total_superadmins = db.execute("SELECT COUNT(*) AS count FROM users WHERE role = 'superadmin'").fetchone()["count"]

    return {
        "total_rooms": total_rooms,
        "active_reservations": active_reservations,
        "active_checkins": active_checkins,
        "occupancy": occupancy,
        "upcoming_checkins": upcoming_checkins,
        "total_users": total_users,
        "total_clients": total_clients,
        "total_admins": total_admins,
        "total_superadmins": total_superadmins,
    }


def parse_month_param(month_str: str | None) -> tuple[date, date, str]:
    if month_str:
        start = datetime.strptime(month_str, "%Y-%m").date().replace(day=1)
    else:
        start = date.today().replace(day=1)

    last_day = calendar.monthrange(start.year, start.month)[1]
    end = start.replace(day=last_day) + timedelta(days=1)
    return start, end, start.strftime("%Y-%m")


def get_monthly_movements(db: sqlite3.Connection, start: date, end: date):
    reservations = db.execute(
        """
        SELECT
            'reservation' AS movement_type,
            r.id AS movement_id,
            r.created_at,
            r.status,
            r.guest_name,
            r.email,
            r.phone,
            r.check_in,
            r.check_out,
            r.guests,
            r.special_requests,
            rm.name AS room_name,
            rm.category AS room_category,
            rm.price_per_night AS price_per_night,
            u.full_name AS user_name,
            u.email AS user_email,
            u.role AS user_role,
            '' AS payment_method,
            0 AS amount_paid,
            0 AS amount_pending
        FROM reservations r
        JOIN rooms rm ON rm.id = r.room_id
        JOIN users u ON u.id = r.user_id
        WHERE date(r.created_at) >= date(?)
          AND date(r.created_at) < date(?)
        """,
        (start.isoformat(), end.isoformat()),
    ).fetchall()

    checkins = db.execute(
        """
        SELECT
            'checkin' AS movement_type,
            hc.id AS movement_id,
            hc.created_at,
            hc.status,
            hc.guest_name,
            '' AS email,
            hc.phone,
            hc.check_in,
            hc.check_out,
            hc.guests,
            hc.special_requests,
            rm.name AS room_name,
            rm.category AS room_category,
            rm.price_per_night AS price_per_night,
            u.full_name AS user_name,
            u.email AS user_email,
            u.role AS user_role,
            hc.payment_method,
            hc.amount_paid,
            hc.amount_pending
        FROM hotel_checkins hc
        JOIN rooms rm ON rm.id = hc.room_id
        JOIN users u ON u.id = hc.created_by
        WHERE date(hc.created_at) >= date(?)
          AND date(hc.created_at) < date(?)
        """,
        (start.isoformat(), end.isoformat()),
    ).fetchall()

    return sorted(list(reservations) + list(checkins), key=lambda x: x["created_at"], reverse=True)


def build_monthly_excel(movements, month_label: str):
    wb = Workbook()

    reservations = [mv for mv in movements if mv["movement_type"] == "reservation"]
    checkins = [mv for mv in movements if mv["movement_type"] == "checkin"]

    headers = [
        "Tipo",
        "ID",
        "Fecha creación",
        "Usuario",
        "Correo usuario",
        "Rol usuario",
        "Huésped",
        "Correo huésped",
        "Teléfono",
        "Habitación",
        "Categoría",
        "Check in",
        "Check out",
        "Noches",
        "Huéspedes",
        "Persona extra",
        "Precio por noche",
        "Método pago",
        "Pagado",
        "Pendiente",
        "Estado",
        "Total",
        "Solicitudes especiales",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    widths = {
        "A": 14, "B": 10, "C": 20, "D": 24, "E": 28, "F": 14, "G": 22, "H": 24,
        "I": 18, "J": 22, "K": 18, "L": 12, "M": 12, "N": 10, "O": 10, "P": 14,
        "Q": 14, "R": 16, "S": 14, "T": 14, "U": 14, "V": 14, "W": 35,
    }

    def fill_sheet(ws, rows, sheet_title: str):
        ws.title = sheet_title
        ws.append(headers)

        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        total_real_checkins = 0.0
        total_estimado_reservas = 0.0
        room_counter: dict[str, int] = {}
        status_counter: dict[str, int] = {}

        for mv in rows:
            check_in_d = datetime.strptime(mv["check_in"], "%Y-%m-%d").date()
            check_out_d = datetime.strptime(mv["check_out"], "%Y-%m-%d").date()
            nights = max((check_out_d - check_in_d).days, 1)
            extra = reservation_has_extra_person_from_text(mv["special_requests"])
            total = calculate_stay_total(
                float(mv["price_per_night"]),
                mv["check_in"],
                mv["check_out"],
                extra
            )

            if mv["movement_type"] == "checkin":
                total_real_checkins += float(mv["amount_paid"] or 0)
            else:
                total_estimado_reservas += total

            room_counter[mv["room_name"]] = room_counter.get(mv["room_name"], 0) + 1
            status_counter[mv["status"]] = status_counter.get(mv["status"], 0) + 1

            ws.append([
                "Check-in" if mv["movement_type"] == "checkin" else "Reservación",
                mv["movement_id"],
                mv["created_at"],
                mv["user_name"],
                mv["user_email"],
                mv["user_role"],
                mv["guest_name"],
                mv["email"],
                mv["phone"],
                mv["room_name"],
                mv["room_category"],
                mv["check_in"],
                mv["check_out"],
                nights,
                mv["guests"],
                "Sí" if extra else "No",
                float(mv["price_per_night"]),
                mv["payment_method"],
                float(mv["amount_paid"] or 0),
                float(mv["amount_pending"] or 0),
                mv["status"],
                total,
                mv["special_requests"],
            ])

        for row in ws.iter_rows(min_row=2, min_col=17, max_col=22):
            for cell in row:
                if cell.column in {17, 19, 20, 22}:
                    cell.number_format = "Q #,##0.00"

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        return {
            "cantidad": len(rows),
            "total_real_checkins": total_real_checkins,
            "total_estimado_reservas": total_estimado_reservas,
            "room_counter": room_counter,
            "status_counter": status_counter,
        }

    # Primera hoja: Reservaciones
    ws_res = wb.active
    res_stats = fill_sheet(ws_res, reservations, "Reservaciones")

    # Segunda hoja: Check-ins
    ws_chk = wb.create_sheet("Checkins")
    chk_stats = fill_sheet(ws_chk, checkins, "Checkins")

    # Hoja resumen
    summary = wb.create_sheet("Resumen")
    summary["A1"] = f"Resumen mensual - {month_label}"
    summary["A1"].font = Font(size=14, bold=True)

    summary["A3"] = "Total reservaciones"
    summary["B3"] = res_stats["cantidad"]

    summary["A4"] = "Total check-ins"
    summary["B4"] = chk_stats["cantidad"]

    summary["A5"] = "Ingreso real por check-ins"
    summary["B5"] = chk_stats["total_real_checkins"]
    summary["B5"].number_format = "Q #,##0.00"

    summary["A6"] = "Valor estimado de reservaciones"
    summary["B6"] = res_stats["total_estimado_reservas"]
    summary["B6"].number_format = "Q #,##0.00"

    summary["A8"] = "Estados reservaciones"
    summary["A8"].font = Font(bold=True)
    r = 9
    for key, value in sorted(res_stats["status_counter"].items()):
        summary[f"A{r}"] = key
        summary[f"B{r}"] = value
        r += 1

    summary["D8"] = "Estados check-ins"
    summary["D8"].font = Font(bold=True)
    r = 9
    for key, value in sorted(chk_stats["status_counter"].items()):
        summary[f"D{r}"] = key
        summary[f"E{r}"] = value
        r += 1

    summary["G8"] = "Habitaciones con más reservaciones"
    summary["G8"].font = Font(bold=True)
    r = 9
    for room_name, qty in sorted(res_stats["room_counter"].items(), key=lambda x: x[1], reverse=True)[:10]:
        summary[f"G{r}"] = room_name
        summary[f"H{r}"] = qty
        r += 1

    summary["J8"] = "Habitaciones con más check-ins"
    summary["J8"].font = Font(bold=True)
    r = 9
    for room_name, qty in sorted(chk_stats["room_counter"].items(), key=lambda x: x[1], reverse=True)[:10]:
        summary[f"J{r}"] = room_name
        summary[f"K{r}"] = qty
        r += 1

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 16
    summary.column_dimensions["D"].width = 24
    summary.column_dimensions["E"].width = 16
    summary.column_dimensions["G"].width = 28
    summary.column_dimensions["H"].width = 14
    summary.column_dimensions["J"].width = 28
    summary.column_dimensions["K"].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================================
# HOME / AUTH
# ==========================================================
@app.get("/")
def index() -> str:
    db = get_db()
    featured_rooms = db.execute(
        """
        SELECT id, name, category, price_per_night, capacity
        FROM rooms
        WHERE capacity > 0
        ORDER BY price_per_night DESC, name ASC
        LIMIT 6
        """
    ).fetchall()
    return render_template("index.html", featured_rooms=featured_rooms)


@app.get("/book")
def book_now() -> str:
    if session.get("user_id"):
        return redirect(url_for("new_reservation"))
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register() -> str:
    if request.method == "POST":
        full_name = (request.form.get("full_name") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""

        if not full_name or not email or not password:
            flash("Completa todos los campos.", "error")
            return render_template("register.html")

        db = get_db()
        exists = db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
        if exists is not None:
            flash("Ese correo ya está registrado. Inicia sesión.", "error")
            return redirect(url_for("login"))

        db.execute(
            """
            INSERT INTO users (full_name, email, password_hash, role)
            VALUES (?, ?, ?, 'client')
            """,
            (full_name, email, generate_password_hash(password)),
        )
        db.commit()
        flash("Cuenta creada. Ya puedes iniciar sesión.", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login() -> str:
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""

        db = get_db()
        user = db.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        if user is None or not check_password_hash(user["password_hash"], password):
            flash("Correo o contraseña incorrectos.", "error")
            return render_template("login.html")

        session.clear()
        session["user_id"] = user["id"]
        flash("Sesión iniciada.", "success")

        if user["role"] == "superadmin":
            return redirect(url_for("superadmin_dashboard"))
        if user["role"] == "admin":
            return redirect(url_for("admin_dashboard"))
        return redirect(url_for("my_reservations"))

    return render_template("login.html")


@app.get("/logout")
def logout() -> str:
    session.clear()
    flash("Sesión cerrada.", "success")
    return redirect(url_for("login"))


# ==========================================================
# CLIENT
# ==========================================================
@app.route("/reservations/new", methods=["GET", "POST"])
@login_required
def new_reservation() -> str:
    db = get_db()
    rooms_meta, categories = get_room_meta(db)

    if request.method == "POST":
        try:
            reservation = parse_reservation_form(request.form)
        except (KeyError, ValueError):
            flash("Hay campos inválidos o incompletos en el formulario.", "error")
            return render_template("reservation_form.html", rooms_meta=rooms_meta, categories=categories, extra_fee=EXTRA_PERSON_FEE)

        errors = validate_reservation(reservation, db)
        if errors:
            for e in errors:
                flash(e, "error")
            return render_template("reservation_form.html", rooms_meta=rooms_meta, categories=categories, extra_fee=EXTRA_PERSON_FEE)

        special = reservation.special_requests
        if reservation.extra_person:
            extra_line = f"[Persona extra: +Q{EXTRA_PERSON_FEE:.2f} por noche]"
            special = (special + "\n" + extra_line).strip() if special else extra_line

        db.execute(
            """
            INSERT INTO reservations
                (user_id, guest_name, email, phone, room_id, check_in, check_out, guests, special_requests, status, admin_alert_dismissed)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'confirmed', 0)
            """,
            (
                session["user_id"],
                reservation.guest_name,
                reservation.email,
                reservation.phone,
                reservation.room_id,
                reservation.check_in.isoformat(),
                reservation.check_out.isoformat(),
                reservation.guests,
                special,
            ),
        )
        db.commit()
        flash("Reservación registrada exitosamente.", "success")
        return redirect(url_for("my_reservations"))

    return render_template("reservation_form.html", rooms_meta=rooms_meta, categories=categories, extra_fee=EXTRA_PERSON_FEE)


@app.route("/my-reservations", endpoint="my_reservations")
@login_required
def my_reservations() -> str:
    user = current_user()
    assert user is not None
    db = get_db()
    reservations = db.execute(
        """
        SELECT r.*, rm.name AS room_name
        FROM reservations r
        JOIN rooms rm ON rm.id = r.room_id
        WHERE r.user_id = ?
        ORDER BY r.created_at DESC
        """,
        (user["id"],),
    ).fetchall()
    return render_template("my_reservations.html", reservations=reservations)


@app.route("/rooms")
def rooms() -> str:
    db = get_db()
    room_list = db.execute("SELECT * FROM rooms WHERE capacity > 0 ORDER BY category, name").fetchall()
    return render_template("rooms.html", rooms=room_list)


# ==========================================================
# ADMIN
# ==========================================================
@app.route("/admin")
@admin_required
def admin_dashboard() -> str:
    db = get_db()
    stats = reservation_stats(db)
    pending_alerts = get_pending_admin_alerts(db)
    has_pending_alerts = len(pending_alerts) > 0
    return render_template("dashboard.html", stats=stats, pending_alerts=pending_alerts, has_pending_alerts=has_pending_alerts)


@app.route("/admin/reservations")
@admin_required
def admin_reservations() -> str:
    db = get_db()
    status = request.args.get("status", "active")
    q = (request.args.get("q") or "").strip().lower()

    base_query = """
        SELECT
            r.*,
            rm.name AS room_name,
            u.full_name AS user_name,
            EXISTS(
                SELECT 1
                FROM hotel_checkins hc
                WHERE hc.reservation_id = r.id
          ) AS has_checkin
        FROM reservations r
        JOIN rooms rm ON rm.id = r.room_id
        JOIN users u ON u.id = r.user_id
        WHERE 1=1
    """
    params: list[Any] = []

    if status == "checked_out":
        base_query += " AND r.status = 'checked_out'"
    elif status == "all":
        pass
    else:
        base_query += " AND r.status != 'checked_out'"

    if q:
        base_query += """
            AND (
                LOWER(r.guest_name) LIKE ?
                OR LOWER(COALESCE(r.phone, '')) LIKE ?
                OR LOWER(rm.name) LIKE ?
                OR LOWER(u.full_name) LIKE ?
                OR LOWER(COALESCE(r.check_in, '')) LIKE ?
                OR LOWER(COALESCE(r.check_out, '')) LIKE ?
                OR LOWER(COALESCE(r.status, '')) LIKE ?
            )
        """
        like_q = f"%{q}%"
        params.extend([like_q, like_q, like_q, like_q, like_q, like_q, like_q])

    if status == "all":
        base_query += """
            ORDER BY
                CASE WHEN r.admin_alert_dismissed = 0 THEN 0 ELSE 1 END,
                r.created_at DESC
        """
    else:
        base_query += """
            ORDER BY
                CASE WHEN r.admin_alert_dismissed = 0 THEN 0 ELSE 1 END,
                r.created_at DESC
        """

    reservations = db.execute(base_query, params).fetchall()

    pending_alerts = get_pending_admin_alerts(db)
    has_pending_alerts = len(pending_alerts) > 0

    return render_template(
        "admin_reservations.html",
        reservations=reservations,
        selected_status=status,
        pending_alerts=pending_alerts,
        has_pending_alerts=has_pending_alerts,
        q=q,
    )


@app.post("/admin/reservations/<int:reservation_id>/status")
@admin_required
def admin_update_reservation_status(reservation_id: int) -> str:
    db = get_db()
    status = request.form.get("status", "")
    if status not in {"confirmed", "checked_in", "checked_out", "cancelled"}:
        flash("Estado inválido.", "error")
        return redirect(url_for("admin_reservations"))

    if status in {"checked_out", "cancelled"}:
        db.execute(
            """
            UPDATE reservations
            SET status = ?, admin_alert_dismissed = 1
            WHERE id = ?
            """,
            (status, reservation_id),
        )
    else:
        db.execute("UPDATE reservations SET status = ? WHERE id = ?", (status, reservation_id))
    db.commit()
    flash("Estado de reservación actualizado.", "success")
    return redirect(url_for("admin_reservations"))


@app.post("/admin/reservations/<int:reservation_id>/dismiss-alert")
@admin_required
def dismiss_reservation_alert(reservation_id: int) -> str:
    db = get_db()
    db.execute("UPDATE reservations SET admin_alert_dismissed = 1 WHERE id = ?", (reservation_id,))
    db.commit()
    flash("Alarma de reservación desactivada.", "success")
    return redirect(url_for("admin_reservations"))


@app.route("/admin/hotel")
@admin_required
def admin_hotel() -> str:
    db = get_db()
    today = date.today().isoformat()
    rooms_meta, categories = get_room_meta(db)

    active_checkins = db.execute(
        """
        SELECT hc.*, rm.name AS room_name, u.full_name AS created_by_name
        FROM hotel_checkins hc
        JOIN rooms rm ON rm.id = hc.room_id
        JOIN users u ON u.id = hc.created_by
        WHERE hc.status = 'checked_in'
        ORDER BY hc.created_at DESC
        """
    ).fetchall()

    checkins_today = db.execute(
        """
        SELECT COUNT(*) AS total
        FROM hotel_checkins
        WHERE status = 'checked_in'
          AND check_in = ?
        """,
        (today,),
    ).fetchone()["total"] or 0

    checkouts_today = db.execute(
        """
        SELECT COUNT(*) AS total
        FROM hotel_checkins
        WHERE status = 'checked_out'
          AND substr(checked_out_at, 1, 10) = ?
        """,
        (today,),
    ).fetchone()["total"] or 0

    occupied_rooms = db.execute(
        """
        SELECT COUNT(DISTINCT room_id) AS total
        FROM hotel_checkins
        WHERE status = 'checked_in'
        """
    ).fetchone()["total"] or 0

    return render_template(
        "admin_hotel.html",
        rooms_meta=rooms_meta,
        categories=categories,
        extra_fee=EXTRA_PERSON_FEE,
        today=today,
        active_checkins=active_checkins,
        checkins_today=checkins_today,
        checkouts_today=checkouts_today,
        occupied_rooms=occupied_rooms,
    )


@app.route("/admin/hotel/checkin", methods=["POST"])
@admin_required
def admin_hotel_checkin() -> str:
    db = get_db()
    rooms_meta, categories = get_room_meta(db)
    today = date.today().isoformat()

    try:
        data = parse_reservation_form(request.form)
    except (KeyError, ValueError):
        flash("Hay campos inválidos o incompletos.", "error")
        return render_template(
            "admin_hotel.html",
            rooms_meta=rooms_meta,
            categories=categories,
            extra_fee=EXTRA_PERSON_FEE,
            today=today,
            active_checkins=[],
            checkins_today=0,
            checkouts_today=0,
            occupied_rooms=0,
        )

    errors = validate_direct_checkin(data, db)
    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("admin_hotel"))

    room = db.execute("SELECT * FROM rooms WHERE id = ?", (data.room_id,)).fetchone()
    assert room is not None

    total_amount = calculate_stay_total(
        float(room["price_per_night"]),
        data.check_in.isoformat(),
        data.check_out.isoformat(),
        data.extra_person
    )

    payment_method = (request.form.get("payment_method") or "cash").strip().lower()
    if payment_method not in {"cash", "card", "transfer"}:
        payment_method = "cash"

    amount_paid = total_amount
    amount_pending = 0.0

    db.execute(
        """
        INSERT INTO hotel_checkins
            (reservation_id, source_type, guest_name, phone, room_id, check_in, check_out,
             guests, extra_person, special_requests, payment_method, total_amount,
             amount_paid, amount_pending, status, created_by)
        VALUES (NULL, 'direct', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'checked_in', ?)
        """,
        (
            data.guest_name,
            data.phone,
            data.room_id,
            data.check_in.isoformat(),
            data.check_out.isoformat(),
            data.guests,
            1 if data.extra_person else 0,
            data.special_requests,
            payment_method,
            total_amount,
            amount_paid,
            amount_pending,
            session["user_id"],
        ),
    )
    db.commit()

    flash("Ingreso registrado correctamente y pagado en su totalidad.", "success")
    return redirect(url_for("admin_hotel_hospedados"))


@app.route("/admin/hotel/hospedados")
@admin_required
def admin_hotel_hospedados() -> str:
    db = get_db()
    today = date.today().isoformat()

    hospedados = db.execute(
        """
        SELECT hc.*, rm.name AS room_name, rm.category AS room_category, u.full_name AS created_by_name
        FROM hotel_checkins hc
        JOIN rooms rm ON rm.id = hc.room_id
        JOIN users u ON u.id = hc.created_by
        WHERE hc.status = 'checked_in'
          AND hc.check_in <= ?
          AND hc.check_out > ?
        ORDER BY hc.check_in DESC, hc.created_at DESC
        """,
        (today, today),
    ).fetchall()

    return render_template("admin_hotel_hospedados.html", hospedados=hospedados, today=today)


@app.route("/admin/hotel/reservations")
@admin_required
def admin_hotel_reservations() -> str:
    db = get_db()
    reservations = db.execute(
        """
        SELECT r.*, rm.name AS room_name, rm.category AS room_category, rm.price_per_night, u.full_name AS user_name
        FROM reservations r
        JOIN rooms rm ON rm.id = r.room_id
        JOIN users u ON u.id = r.user_id
        WHERE r.status IN ('confirmed', 'checked_in')
        ORDER BY r.check_in ASC, r.created_at DESC
        """
    ).fetchall()
    return render_template("admin_hotel_reservations.html", reservations=reservations)


@app.route("/admin/hotel/reservations/<int:reservation_id>/checkin", methods=["GET", "POST"])
@admin_required
def admin_hotel_reservation_to_checkin(reservation_id: int) -> str:
    db = get_db()
    reservation = db.execute(
        """
        SELECT r.*, rm.name AS room_name, rm.category AS room_category, rm.price_per_night, rm.capacity
        FROM reservations r
        JOIN rooms rm ON rm.id = r.room_id
        WHERE r.id = ?
        """,
        (reservation_id,),
    ).fetchone()

    if reservation is None:
        flash("Reservación no encontrada.", "error")
        return redirect(url_for("admin_hotel_reservations"))

    existing_checkin = db.execute(
        "SELECT id FROM hotel_checkins WHERE reservation_id = ? LIMIT 1",
        (reservation_id,),
    ).fetchone()

    if existing_checkin:
        flash("Esta reservación ya fue convertida a check-in.", "error")
        return redirect(url_for("admin_hotel_reservations"))

    if request.method == "POST":
        guest_name = (request.form.get("guest_name") or reservation["guest_name"]).strip()
        phone = (request.form.get("phone") or reservation["phone"]).strip()
        check_in = request.form.get("check_in") or reservation["check_in"]
        check_out = request.form.get("check_out") or reservation["check_out"]
        guests = int(request.form.get("guests") or reservation["guests"] or 1)
        extra_person = 1 if request.form.get("extra_person") == "1" else 0
        special_requests = (request.form.get("special_requests") or reservation["special_requests"] or "").strip()

        payment_method = (request.form.get("payment_method") or "cash").strip().lower()
        if payment_method not in {"cash", "card", "transfer"}:
            payment_method = "cash"

        total_amount = calculate_stay_total(
            float(reservation["price_per_night"]),
            check_in,
            check_out,
            bool(extra_person),
        )
        amount_paid = total_amount
        amount_pending = 0.0

        db.execute(
            """
            INSERT INTO hotel_checkins
                (reservation_id, source_type, guest_name, phone, room_id, check_in, check_out,
                 guests, extra_person, special_requests, payment_method, total_amount,
                 amount_paid, amount_pending, status, created_by)
            VALUES (?, 'reservation', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'checked_in', ?)
            """,
            (
                reservation_id,
                guest_name,
                phone,
                reservation["room_id"],
                check_in,
                check_out,
                guests,
                extra_person,
                special_requests,
                payment_method,
                total_amount,
                amount_paid,
                amount_pending,
                session["user_id"],
            ),
        )

        db.execute(
            "UPDATE reservations SET status = 'checked_in', admin_alert_dismissed = 1 WHERE id = ?",
            (reservation_id,),
        )
        db.commit()

        flash("Reservación convertida a check-in y pagada completamente.", "success")
        return redirect(url_for("admin_hotel_hospedados"))

    return render_template(
        "admin_reservation_to_checkin.html",
        reservation=reservation,
        extra_fee=EXTRA_PERSON_FEE,
    )

@app.post("/admin/hotel/checkins/<int:checkin_id>/checkout")
@admin_required
def admin_hotel_checkout(checkin_id: int) -> str:
    db = get_db()
    row = db.execute("SELECT * FROM hotel_checkins WHERE id = ?", (checkin_id,)).fetchone()

    if row is None:
        flash("Ingreso no encontrado.", "error")
        return redirect(url_for("admin_hotel_hospedados"))

    if row["status"] != "checked_in":
        flash("Este huésped ya no está activo.", "error")
        return redirect(url_for("admin_hotel_hospedados"))

    db.execute(
        """
        UPDATE hotel_checkins
        SET status = 'checked_out',
            checked_out_at = datetime('now'),
            amount_pending = CASE WHEN amount_pending < 0 THEN 0 ELSE amount_pending END
        WHERE id = ?
        """,
        (checkin_id,),
    )
    db.commit()

    flash("Check-out realizado correctamente.", "success")
    return redirect(url_for("admin_hotel_hospedados"))


# ==========================================================
# SUPERADMIN
# ==========================================================
@app.route("/superadmin")
@superadmin_required
def superadmin_dashboard() -> str:
    db = get_db()
    stats = reservation_stats(db)
    start, end, month_label = parse_month_param(request.args.get("month"))
    movements = get_monthly_movements(db, start, end)

    monthly_income = sum(float(mv["amount_paid"] or 0) for mv in movements if mv["movement_type"] == "checkin")
    pending_alerts = get_pending_admin_alerts(db)
    has_pending_alerts = len(pending_alerts) > 0

    top_rooms: dict[str, int] = {}
    status_counter: dict[str, int] = {
        "confirmed": 0,
        "checked_in": 0,
        "checked_out": 0,
        "cancelled": 0,
    }
    ingresos_por_dia: dict[str, float] = {}

    for mv in movements:
        top_rooms[mv["room_name"]] = top_rooms.get(mv["room_name"], 0) + 1

        status = mv["status"] or ""
        status_counter[status] = status_counter.get(status, 0) + 1

        fecha = str(mv["created_at"])[:10]
        valor = float(mv["amount_paid"] or 0) if mv["movement_type"] == "checkin" else 0.0
        ingresos_por_dia[fecha] = ingresos_por_dia.get(fecha, 0.0) + valor

    top_rooms_sorted = sorted(top_rooms.items(), key=lambda x: x[1], reverse=True)[:6]
    ingresos_por_dia_sorted = sorted(ingresos_por_dia.items(), key=lambda x: x[0])

    return render_template(
        "superadmin_dashboard.html",
        stats=stats,
        month_label=month_label,
        monthly_movements=movements,
        monthly_income=monthly_income,
        top_rooms=top_rooms_sorted,
        pending_alerts=pending_alerts,
        has_pending_alerts=has_pending_alerts,
        status_counter=status_counter,
        ingresos_por_dia=ingresos_por_dia_sorted,
        reservation_total_for_template=calculate_reservation_total,
    )

@app.route("/superadmin/users", methods=["GET", "POST"])
@superadmin_required
def superadmin_users() -> str:
    db = get_db()

    if request.method == "POST":
        full_name = (request.form.get("full_name") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        role = (request.form.get("role") or "client").strip()

        if role not in {"client", "admin", "superadmin"}:
            flash("Rol inválido.", "error")
            return redirect(url_for("superadmin_users"))

        if not full_name or not email or not password:
            flash("Debes completar nombre, correo y contraseña.", "error")
            return redirect(url_for("superadmin_users"))

        exists = db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
        if exists:
            flash("Ese correo ya existe.", "error")
            return redirect(url_for("superadmin_users"))

        db.execute(
            """
            INSERT INTO users (full_name, email, password_hash, role)
            VALUES (?, ?, ?, ?)
            """,
            (full_name, email, generate_password_hash(password), role),
        )
        db.commit()
        flash("Usuario creado correctamente.", "success")
        return redirect(url_for("superadmin_users"))

    q = (request.args.get("q") or "").strip().lower()
    role_filter = (request.args.get("role") or "").strip()

    sql = """
        SELECT u.*, COUNT(r.id) AS total_reservations
        FROM users u
        LEFT JOIN reservations r ON r.user_id = u.id
        WHERE 1=1
    """
    params: list[Any] = []

    if q:
        sql += " AND (LOWER(u.full_name) LIKE ? OR LOWER(u.email) LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%"])

    if role_filter in {"client", "admin", "superadmin"}:
        sql += " AND u.role = ?"
        params.append(role_filter)

    sql += " GROUP BY u.id ORDER BY u.created_at DESC"
    users = db.execute(sql, params).fetchall()

    return render_template("superadmin_users.html", users=users, q=q, role_filter=role_filter)


@app.post("/superadmin/users/<int:user_id>/role")
@superadmin_required
def superadmin_update_user_role(user_id: int) -> str:
    db = get_db()
    new_role = (request.form.get("role") or "").strip()

    if new_role not in {"client", "admin", "superadmin"}:
        flash("Rol inválido.", "error")
        return redirect(url_for("superadmin_users"))

    current = current_user()
    target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if target is None:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("superadmin_users"))

    if current and current["id"] == user_id and new_role != "superadmin":
        flash("No puedes quitarte tu propio rol de superadmin.", "error")
        return redirect(url_for("superadmin_users"))

    db.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, user_id))
    db.commit()
    flash("Rol actualizado correctamente.", "success")
    return redirect(url_for("superadmin_users"))


@app.post("/superadmin/users/<int:user_id>/reset-password")
@superadmin_required
def superadmin_reset_password(user_id: int) -> str:
    db = get_db()
    new_password = request.form.get("new_password") or ""

    if len(new_password) < 6:
        flash("La nueva contraseña debe tener al menos 6 caracteres.", "error")
        return redirect(url_for("superadmin_users"))

    target = db.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
    if target is None:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("superadmin_users"))

    db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), user_id))
    db.commit()
    flash("Contraseña restablecida correctamente.", "success")
    return redirect(url_for("superadmin_users"))


@app.post("/superadmin/users/<int:user_id>/delete")
@superadmin_required
def superadmin_delete_user(user_id: int) -> str:
    db = get_db()
    current = current_user()
    target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()

    if target is None:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("superadmin_users"))

    if current and current["id"] == user_id:
        flash("No puedes eliminar tu propio usuario.", "error")
        return redirect(url_for("superadmin_users"))

    reservations_count = db.execute("SELECT COUNT(*) AS count FROM reservations WHERE user_id = ?", (user_id,)).fetchone()["count"]
    if reservations_count > 0:
        flash("No puedes eliminar este usuario porque tiene reservaciones asociadas.", "error")
        return redirect(url_for("superadmin_users"))

    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    flash("Usuario eliminado correctamente.", "success")
    return redirect(url_for("superadmin_users"))


@app.get("/superadmin/export/monthly-report")
@superadmin_required
def superadmin_export_monthly_report():
    db = get_db()
    start, end, month_label = parse_month_param(request.args.get("month"))
    movements = get_monthly_movements(db, start, end)
    excel_file = build_monthly_excel(movements, month_label)
    filename = f"movimiento_mensual_{month_label}.xlsx"

    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ==========================================================
# MAIN
# ==========================================================
if __name__ == "__main__":
    if not DB_PATH.exists():
        init_db()

    ensure_reservations_alert_column()
    ensure_role_indexes_and_superadmin()
    ensure_hotel_checkins_table()

    res = apply_room_updates()
    print(f"[ROOMS] actualizadas: {res['updated']} | renombres: {res['renamed']}")

    hidden = hide_non_official_rooms()
    print(f"[ROOMS] ocultadas (no oficiales): {hidden}")

    app.run(host="0.0.0.0", port=5000, debug=True)